import copy
import re
import zipfile
import shutil
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
import structlog

logger = structlog.get_logger()

BALANCE_SHEET_NAME = "Balance Sheet"
DATA_START_ROW = 11
INSERT_COL = 2
INSERT_COUNT = 3

SECTION_NORMALIZE = {
    "current assets": "current_assets",
    "non current assets": "noncurrent_assets",
    "noncurrent assets": "noncurrent_assets",
    "non-current assets": "noncurrent_assets",
    "current liabilities": "current_liabilities",
    "non-current liabilities": "noncurrent_liabilities",
    "noncurrent liabilities": "noncurrent_liabilities",
    "non current liabilities": "noncurrent_liabilities",
    "equity": "equity",
    "commitments and contingencies": "commitments",
    "commitments & contingencies": "commitments",
}


class TemplatePopulationService:
    def process(
        self,
        template_path: str,
        extraction_data: dict,
        output_path: str,
    ) -> str:
        wb = load_workbook(template_path, keep_vba=True)

        if BALANCE_SHEET_NAME not in wb.sheetnames:
            wb.close()
            raise ValueError(f"Sheet '{BALANCE_SHEET_NAME}' not found in template")

        ws = wb[BALANCE_SHEET_NAME]

        period, values_by_section_category = self._prepare_data(extraction_data)
        logger.info(
            "template_population_started",
            period=period,
            categories=len(values_by_section_category),
        )

        header_merges = self._capture_header_merges(ws)

        ws.insert_cols(INSERT_COL, amount=INSERT_COUNT)
        self._fix_shifted_formulas(ws)
        self._fix_cross_sheet_references(wb, BALANCE_SHEET_NAME, INSERT_COL, INSERT_COUNT)
        self._fix_header_merges(ws, header_merges)

        self._style_new_columns(ws)
        self._set_new_column_headers(ws, period)
        populated = self._populate_values(ws, values_by_section_category)
        self._propagate_total_formulas(ws)

        wb.save(output_path)
        wb.close()

        self._restore_external_links(template_path, output_path)

        logger.info("template_population_completed", output=output_path, rows_populated=populated)
        return output_path

    def _prepare_data(self, extraction_data: dict) -> tuple[str, dict]:
        periods = extraction_data.get("periods", [])
        rows = extraction_data.get("rows", [])

        if not periods:
            raise ValueError("No periods found in extraction data")

        most_recent_period = periods[0]

        values_by_section_category = {}
        for row in rows:
            category = row.get("mapped_category", "")
            if not category or category == "UNMAPPED" or row.get("is_non_mappable"):
                continue
            section = row.get("section", "")
            norm_section = self._normalize_section(section)
            value = row.get("values", {}).get(most_recent_period)
            if value is not None:
                values_by_section_category[(norm_section, category)] = value

        return most_recent_period, values_by_section_category

    def _normalize_section(self, section: str) -> str:
        return SECTION_NORMALIZE.get(section.lower().strip(), section.lower().strip())

    def _capture_header_merges(self, ws) -> list:
        merges = []
        for m in ws.merged_cells.ranges:
            if m.min_row <= 9:
                merges.append({
                    "min_row": m.min_row,
                    "max_row": m.max_row,
                    "min_col": m.min_col,
                    "max_col": m.max_col,
                })
        return merges

    def _fix_header_merges(self, ws, original_merges: list):
        for m in list(ws.merged_cells.ranges):
            if m.min_row <= 9:
                try:
                    ws.unmerge_cells(str(m))
                except (KeyError, ValueError):
                    pass

        for m in original_merges:
            min_col = m["min_col"]
            max_col = m["max_col"]
            if min_col >= INSERT_COL:
                min_col += INSERT_COUNT
                max_col += INSERT_COUNT
            try:
                ws.merge_cells(
                    start_row=m["min_row"],
                    start_column=min_col,
                    end_row=m["max_row"],
                    end_column=max_col,
                )
            except (ValueError, KeyError):
                pass

        src_col = INSERT_COL + INSERT_COUNT
        for m in list(ws.merged_cells.ranges):
            if m.min_col == src_col and m.max_col == src_col + 1 and m.min_row <= 9:
                try:
                    ws.merge_cells(
                        start_row=m.min_row,
                        start_column=INSERT_COL,
                        end_row=m.max_row,
                        end_column=INSERT_COL + 1,
                    )
                except (ValueError, KeyError):
                    pass

        for m in list(ws.merged_cells.ranges):
            if m.min_col == src_col + 2 and m.min_row <= 9 and m.min_col == m.max_col:
                try:
                    ws.merge_cells(
                        start_row=m.min_row,
                        start_column=INSERT_COL + 2,
                        end_row=m.max_row,
                        end_column=INSERT_COL + 2,
                    )
                except (ValueError, KeyError):
                    pass

    def _fix_shifted_formulas(self, ws):
        shifted_start = INSERT_COL + INSERT_COUNT

        for row in ws.iter_rows(min_col=shifted_start, max_col=ws.max_column,
                                min_row=1, max_row=ws.max_row):
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                    cell.value = self._shift_column_refs(cell.value)

        logger.info("formulas_fixed")

    def _shift_column_refs(self, formula: str) -> str:
        def replace_ref(match):
            prefix = match.group(1)
            dollar_col = match.group(2)
            col_str = match.group(3)
            dollar_row = match.group(4)
            row_num = match.group(5)

            col_idx = column_index_from_string(col_str)
            if col_idx >= INSERT_COL:
                col_idx += INSERT_COUNT
                col_str = get_column_letter(col_idx)

            return f"{prefix}{dollar_col}{col_str}{dollar_row}{row_num}"

        return re.sub(
            r"((?:^|[=,+\-*/&<>!( :]))(\$?)([A-Z]{1,3})(\$?)(\d+)",
            replace_ref,
            formula,
        )

    def _fix_cross_sheet_references(self, wb, modified_sheet: str, insert_col: int, insert_count: int):
        escaped_name = re.escape(modified_sheet)
        pattern = re.compile(
            r"('{0}'!|{0}!)(\$?)([A-Z]{{1,3}})(\$?)(\d+)".format(escaped_name)
        )

        def shift_match(match):
            sheet_ref = match.group(1)
            dollar_col = match.group(2)
            col_str = match.group(3)
            dollar_row = match.group(4)
            row_num = match.group(5)

            col_idx = column_index_from_string(col_str)
            if col_idx >= insert_col:
                col_idx += insert_count
                col_str = get_column_letter(col_idx)

            return f"{sheet_ref}{dollar_col}{col_str}{dollar_row}{row_num}"

        fixed_count = 0
        for sheet_name in wb.sheetnames:
            if sheet_name == modified_sheet:
                continue
            ws = wb[sheet_name]
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row,
                                    min_col=1, max_col=ws.max_column):
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                        new_val = pattern.sub(shift_match, cell.value)
                        if new_val != cell.value:
                            cell.value = new_val
                            fixed_count += 1

        logger.info("cross_sheet_references_fixed", modified_sheet=modified_sheet, refs_updated=fixed_count)

    def _style_new_columns(self, ws):
        src_as_given = INSERT_COL + INSERT_COUNT
        src_as_allowed = src_as_given + 1
        src_remarks = src_as_given + 2

        for row_idx in range(1, ws.max_row + 1):
            self._copy_cell_style(ws, row_idx, src_as_given, row_idx, INSERT_COL)
            self._copy_cell_style(ws, row_idx, src_as_allowed, row_idx, INSERT_COL + 1)
            self._copy_cell_style(ws, row_idx, src_remarks, row_idx, INSERT_COL + 2)

    def _set_new_column_headers(self, ws, period: str):
        period_date = self._parse_period_date(period)
        ws.cell(row=6, column=INSERT_COL, value=period_date)
        ws.cell(row=9, column=INSERT_COL, value="As Given")
        ws.cell(row=9, column=INSERT_COL + 1, value="As Allowed")
        ws.cell(row=6, column=INSERT_COL + 2, value="Remarks")

        src_col = INSERT_COL + INSERT_COUNT
        src_agg = ws.cell(row=7, column=src_col).value
        if src_agg:
            ws.cell(row=7, column=INSERT_COL, value=src_agg)

        period_date = self._parse_period_date(period)
        is_audit = period_date.month == 12 if isinstance(period_date, datetime) else False
        ws.cell(row=8, column=INSERT_COL, value="Audit" if is_audit else "Internal")

    def _populate_values(self, ws, values_by_section_category: dict) -> int:
        populated = 0
        as_given_letter = get_column_letter(INSERT_COL)
        current_section = ""

        for row_idx in range(DATA_START_ROW, ws.max_row + 1):
            label_cell = ws.cell(row=row_idx, column=1)
            label = label_cell.value
            if not label:
                continue

            label_clean = str(label).strip()
            norm = self._normalize_section(label_clean)
            if norm in SECTION_NORMALIZE.values():
                current_section = norm
                continue

            if label_clean.lower().startswith("total"):
                continue

            key = (current_section, label_clean)
            if key in values_by_section_category:
                value = values_by_section_category[key]
                scaled_value = value * 1000
                ws.cell(row=row_idx, column=INSERT_COL, value=scaled_value)
                ws.cell(
                    row=row_idx,
                    column=INSERT_COL + 1,
                    value=f"={as_given_letter}{row_idx}*1",
                )
                populated += 1
                logger.debug(
                    "value_populated",
                    row=row_idx,
                    section=current_section,
                    label=label_clean,
                    value=value,
                )

        return populated

    def _propagate_total_formulas(self, ws):
        src_as_given = INSERT_COL + INSERT_COUNT
        src_as_allowed = src_as_given + 1
        new_as_given_letter = get_column_letter(INSERT_COL)
        new_as_allowed_letter = get_column_letter(INSERT_COL + 1)
        src_as_given_letter = get_column_letter(src_as_given)
        src_as_allowed_letter = get_column_letter(src_as_allowed)

        total_count = 0
        for row_idx in range(DATA_START_ROW, ws.max_row + 1):
            label = ws.cell(row=row_idx, column=1).value
            if not label:
                continue

            label_clean = str(label).strip()
            if not label_clean.lower().startswith("total") and label_clean != "Working Capital":
                continue

            src_given_cell = ws.cell(row=row_idx, column=src_as_given)
            if src_given_cell.value and isinstance(src_given_cell.value, str) and src_given_cell.value.startswith("="):
                new_formula = src_given_cell.value.replace(
                    src_as_given_letter, new_as_given_letter
                )
                ws.cell(row=row_idx, column=INSERT_COL, value=new_formula)

            src_allowed_cell = ws.cell(row=row_idx, column=src_as_allowed)
            if src_allowed_cell.value and isinstance(src_allowed_cell.value, str) and src_allowed_cell.value.startswith("="):
                new_formula = src_allowed_cell.value.replace(
                    src_as_allowed_letter, new_as_allowed_letter
                )
                ws.cell(row=row_idx, column=INSERT_COL + 1, value=new_formula)

            total_count += 1

        logger.info("total_formulas_propagated", count=total_count)

    def _copy_cell_style(self, ws, src_row, src_col, dst_row, dst_col):
        src = ws.cell(row=src_row, column=src_col)
        dst = ws.cell(row=dst_row, column=dst_col)
        if src.has_style:
            dst.font = copy.copy(src.font)
            dst.fill = copy.copy(src.fill)
            dst.border = copy.copy(src.border)
            dst.alignment = copy.copy(src.alignment)
            dst.number_format = src.number_format

    def _parse_period_date(self, period_str: str):
        try:
            return datetime.strptime(period_str, "%B %d, %Y")
        except ValueError:
            return period_str

    def _restore_external_links(self, original_path: str, output_path: str):
        external_files = []
        with zipfile.ZipFile(original_path, "r") as orig_zip:
            for name in orig_zip.namelist():
                if "externalLink" in name:
                    external_files.append(name)

        if not external_files:
            return

        temp_path = output_path + ".tmp"
        with zipfile.ZipFile(output_path, "r") as out_zip:
            with zipfile.ZipFile(temp_path, "w", zipfile.ZIP_DEFLATED) as new_zip:
                for item in out_zip.namelist():
                    if item in external_files:
                        continue
                    new_zip.writestr(item, out_zip.read(item))

                with zipfile.ZipFile(original_path, "r") as orig_zip:
                    for ef in external_files:
                        new_zip.writestr(ef, orig_zip.read(ef))

        shutil.move(temp_path, output_path)
        logger.info("external_links_restored", files=len(external_files))
