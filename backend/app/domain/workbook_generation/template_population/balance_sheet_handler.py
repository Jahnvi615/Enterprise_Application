from datetime import datetime
from openpyxl.utils import get_column_letter
from app.domain.business_rules.service import SpreadingRulesService
import structlog

logger = structlog.get_logger()

BALANCE_SHEET_NAME = "Balance Sheet"
DATA_START_ROW = 11
INSERT_COL = 2
INSERT_COUNT = 3

SECTION_NORMALIZE = {
    "current assets": "current_assets",
    "non current assets": "noncurrent_assets",
    "noncurrent assets": "noncurrent_assets",
    "non-current assets": "noncurrent_assets",
    "current liabilities": "current_liabilities",
    "non-current liabilities": "noncurrent_liabilities",
    "noncurrent liabilities": "noncurrent_liabilities",
    "non current liabilities": "noncurrent_liabilities",
    "equity": "equity",
    "commitments and contingencies": "commitments",
    "commitments & contingencies": "commitments",
}


class BalanceSheetHandler:
    def __init__(self, orchestrator):
        self._orch = orchestrator

    def handle(self, wb, extraction_data: dict, context: dict):
        ws = wb[BALANCE_SHEET_NAME]

        period = context["period"]
        values_by_section_category = self._prepare_data(extraction_data)

        logger.info(
            "balance_sheet_handler_started",
            period=period,
            categories=len(values_by_section_category),
        )

        header_merges = self._orch.capture_header_merges(ws)

        ws.insert_cols(INSERT_COL, amount=INSERT_COUNT)
        self._orch.fix_shifted_formulas(ws, INSERT_COL, INSERT_COUNT)
        self._orch.fix_header_merges(ws, header_merges, INSERT_COL, INSERT_COUNT)

        self._create_new_column_merges(ws)
        self._style_new_columns(ws)
        self._set_new_column_headers(ws, period)
        populated = self._populate_values(ws, values_by_section_category)

        spreading_service = SpreadingRulesService()
        spreading_service.apply(
            ws,
            as_given_col=INSERT_COL,
            as_allowed_col=INSERT_COL + 1,
            remarks_col=INSERT_COL + 2,
        )

        self._propagate_total_formulas(ws)

        logger.info("balance_sheet_handler_completed", rows_populated=populated)

    def _prepare_data(self, extraction_data: dict) -> dict:
        periods = extraction_data.get("periods", [])
        rows = extraction_data.get("rows", [])

        most_recent_period = periods[0]

        values_by_section_category = {}
        for row in rows:
            category = row.get("mapped_category", "")
            if not category or category == "UNMAPPED" or row.get("is_non_mappable"):
                continue
            section = row.get("section", "")
            norm_section = self._normalize_section(section)
            value = row.get("values", {}).get(most_recent_period)
            if value is not None:
                values_by_section_category[(norm_section, category)] = value

        return values_by_section_category

    def _normalize_section(self, section: str) -> str:
        return SECTION_NORMALIZE.get(section.lower().strip(), section.lower().strip())

    def _create_new_column_merges(self, ws):
        src_col = INSERT_COL + INSERT_COUNT

        src_merges = []
        for m in list(ws.merged_cells.ranges):
            if m.min_col == src_col and m.max_col == src_col + 1 and m.min_row <= 9:
                src_merges.append({"min_row": m.min_row, "max_row": m.max_row})

        for m in src_merges:
            try:
                ws.merge_cells(
                    start_row=m["min_row"],
                    start_column=INSERT_COL,
                    end_row=m["max_row"],
                    end_column=INSERT_COL + 1,
                )
            except (ValueError, KeyError):
                pass

        for m in list(ws.merged_cells.ranges):
            if m.min_col == src_col + 2 and m.min_row <= 9 and m.min_col == m.max_col:
                try:
                    ws.merge_cells(
                        start_row=m.min_row,
                        start_column=INSERT_COL + 2,
                        end_row=m.max_row,
                        end_column=INSERT_COL + 2,
                    )
                except (ValueError, KeyError):
                    pass

    def _style_new_columns(self, ws):
        src_as_given = INSERT_COL + INSERT_COUNT
        src_as_allowed = src_as_given + 1
        src_remarks = src_as_given + 2

        for row_idx in range(1, ws.max_row + 1):
            self._orch.copy_cell_style(ws, row_idx, src_as_given, row_idx, INSERT_COL)
            self._orch.copy_cell_style(ws, row_idx, src_as_allowed, row_idx, INSERT_COL + 1)
            self._orch.copy_cell_style(ws, row_idx, src_remarks, row_idx, INSERT_COL + 2)

    def _set_new_column_headers(self, ws, period: str):
        period_date = self._orch.parse_period_date(period)
        ws.cell(row=6, column=INSERT_COL, value=period_date)
        ws.cell(row=9, column=INSERT_COL, value="As Given")
        ws.cell(row=9, column=INSERT_COL + 1, value="As Allowed")
        ws.cell(row=6, column=INSERT_COL + 2, value="Remarks")

        src_col = INSERT_COL + INSERT_COUNT
        src_agg = ws.cell(row=7, column=src_col).value
        if src_agg:
            ws.cell(row=7, column=INSERT_COL, value=src_agg)

        is_audit = period_date.month == 12 if isinstance(period_date, datetime) else False
        ws.cell(row=8, column=INSERT_COL, value="Audit" if is_audit else "Internal")

    def _populate_values(self, ws, values_by_section_category: dict) -> int:
        populated = 0
        current_section = ""

        for row_idx in range(DATA_START_ROW, ws.max_row + 1):
            label_cell = ws.cell(row=row_idx, column=1)
            label = label_cell.value
            if not label:
                continue

            label_clean = str(label).strip()
            norm = self._normalize_section(label_clean)
            if norm in SECTION_NORMALIZE.values():
                current_section = norm
                continue

            if label_clean.lower().startswith("total"):
                continue

            key = (current_section, label_clean)
            if key in values_by_section_category:
                value = values_by_section_category[key]
                scaled_value = value * 1000
                ws.cell(row=row_idx, column=INSERT_COL, value=scaled_value)
                populated += 1
                logger.debug(
                    "value_populated",
                    row=row_idx,
                    section=current_section,
                    label=label_clean,
                    value=value,
                )

        return populated

    def _propagate_total_formulas(self, ws):
        src_as_given = INSERT_COL + INSERT_COUNT
        src_as_allowed = src_as_given + 1
        new_as_given_letter = get_column_letter(INSERT_COL)
        new_as_allowed_letter = get_column_letter(INSERT_COL + 1)
        src_as_given_letter = get_column_letter(src_as_given)
        src_as_allowed_letter = get_column_letter(src_as_allowed)

        total_count = 0
        for row_idx in range(DATA_START_ROW, ws.max_row + 1):
            label = ws.cell(row=row_idx, column=1).value
            if not label:
                continue

            label_clean = str(label).strip()
            if not label_clean.lower().startswith("total") and label_clean != "Working Capital":
                continue

            src_given_cell = ws.cell(row=row_idx, column=src_as_given)
            if src_given_cell.value and isinstance(src_given_cell.value, str) and src_given_cell.value.startswith("="):
                new_formula = src_given_cell.value.replace(
                    src_as_given_letter, new_as_given_letter
                )
                ws.cell(row=row_idx, column=INSERT_COL, value=new_formula)

            src_allowed_cell = ws.cell(row=row_idx, column=src_as_allowed)
            if src_allowed_cell.value and isinstance(src_allowed_cell.value, str) and src_allowed_cell.value.startswith("="):
                new_formula = src_allowed_cell.value.replace(
                    src_as_allowed_letter, new_as_allowed_letter
                )
                ws.cell(row=row_idx, column=INSERT_COL + 1, value=new_formula)

            total_count += 1

        logger.info("total_formulas_propagated", count=total_count)
