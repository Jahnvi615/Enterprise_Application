import copy
import re
import zipfile
import shutil
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.workbook.properties import CalcProperties
from openpyxl.utils import get_column_letter, column_index_from_string
from app.domain.workbook_generation.template_population.balance_sheet_handler import BalanceSheetHandler
from app.domain.workbook_generation.template_population.balance_sheet_trend_handler import BalanceSheetTrendHandler
from app.domain.workbook_generation.template_population.ratio_handler import RatioHandler
from app.domain.workbook_generation.template_population.cash_flow_handler import CashFlowHandler
from app.domain.workbook_generation.template_population.p_and_l_handler import PAndLHandler, PNL_SHEET_NAME
import structlog

logger = structlog.get_logger()

BALANCE_SHEET_NAME = "Balance Sheet"
BALANCE_SHEET_TREND_NAME = "Balance Sheet Trend"


class TemplatePopulationService:
    def process(
        self,
        template_path: str,
        extraction_data: dict,
        output_path: str,
        cash_flow_data: dict | None = None,
        p_and_l_data: dict | None = None,
        notes_data: dict | None = None,
    ) -> str:
        wb = load_workbook(template_path, keep_vba=True)

        period = self._extract_period(extraction_data)
        period_date = self._parse_period_date(period)

        context = {
            "period": period,
            "period_date": period_date,
            "insert_col": 2,
            "bs_insert_count": 3,
            "bs_as_allowed_col": 3,
        }

        logger.info("template_population_started", period=period)

        if BALANCE_SHEET_NAME in wb.sheetnames:
            bs_handler = BalanceSheetHandler(self)
            bs_handler.handle(wb, extraction_data, context, notes_data=notes_data or {})
            self.fix_cross_sheet_references(wb, BALANCE_SHEET_NAME, context["insert_col"], context["bs_insert_count"])

        if BALANCE_SHEET_TREND_NAME in wb.sheetnames:
            trend_handler = BalanceSheetTrendHandler(self)
            trend_handler.handle(wb, context)
            self.fix_cross_sheet_references(wb, BALANCE_SHEET_TREND_NAME, 2, 1)

        if PNL_SHEET_NAME in wb.sheetnames and p_and_l_data:
            pnl_handler = PAndLHandler(self)
            pnl_handler.handle(wb, p_and_l_data, context)
            self.fix_cross_sheet_references(wb, PNL_SHEET_NAME, 2, 1)

        if "Ratio" in wb.sheetnames:
            ratio_handler = RatioHandler(self)
            ratio_handler.handle(wb, context)
            self.fix_cross_sheet_references(wb, "Ratio", 4, 1)

        if "Cash Flow" in wb.sheetnames and cash_flow_data:
            cf_handler = CashFlowHandler(self)
            cf_handler.handle(wb, cash_flow_data, context)
            self.fix_cross_sheet_references(wb, "Cash Flow", 2, 1)

        wb.calculation = CalcProperties(fullCalcOnLoad=True)
        wb.save(output_path)
        wb.close()

        self._restore_external_links(template_path, output_path)

        logger.info("template_population_completed", output=output_path)
        return output_path

    def _extract_period(self, extraction_data: dict) -> str:
        periods = extraction_data.get("periods", [])
        if not periods:
            raise ValueError("No periods found in extraction data")
        return periods[0]

    # --- Shared Utilities ---

    def copy_cell_style(self, ws, src_row, src_col, dst_row, dst_col):
        src = ws.cell(row=src_row, column=src_col)
        dst = ws.cell(row=dst_row, column=dst_col)
        if src.has_style:
            dst.font = copy.copy(src.font)
            dst.fill = copy.copy(src.fill)
            dst.border = copy.copy(src.border)
            dst.alignment = copy.copy(src.alignment)
            dst.number_format = src.number_format

    def fix_shifted_formulas(self, ws, insert_col, insert_count):
        shifted_start = insert_col + insert_count

        for row in ws.iter_rows(min_col=shifted_start, max_col=ws.max_column,
                                min_row=1, max_row=ws.max_row):
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                    cell.value = self.shift_column_refs(cell.value, insert_col, insert_count)

    def shift_column_refs(self, formula: str, insert_col: int, insert_count: int) -> str:
        def replace_ref(match):
            prefix = match.group(1)
            dollar_col = match.group(2)
            col_str = match.group(3)
            dollar_row = match.group(4)
            row_num = match.group(5)

            col_idx = column_index_from_string(col_str)
            if col_idx >= insert_col:
                col_idx += insert_count
                col_str = get_column_letter(col_idx)

            return f"{prefix}{dollar_col}{col_str}{dollar_row}{row_num}"

        return re.sub(
            r"((?:^|[=,+\-*/&<>( :]))(\$?)([A-Z]{1,3})(\$?)(\d+)",
            replace_ref,
            formula,
        )

    def fix_cross_sheet_references(self, wb, modified_sheet: str, insert_col: int, insert_count: int):
        escaped_name = re.escape(modified_sheet)
        pattern = re.compile(
            r"('{0}'!|{0}!)(\$?)([A-Z]{{1,3}})(\$?)(\d+)".format(escaped_name)
        )

        def shift_match(match):
            sheet_ref = match.group(1)
            dollar_col = match.group(2)
            col_str = match.group(3)
            dollar_row = match.group(4)
            row_num = match.group(5)

            col_idx = column_index_from_string(col_str)
            if col_idx >= insert_col:
                col_idx += insert_count
                col_str = get_column_letter(col_idx)

            return f"{sheet_ref}{dollar_col}{col_str}{dollar_row}{row_num}"

        fixed_count = 0
        for sheet_name in wb.sheetnames:
            if sheet_name == modified_sheet:
                continue
            ws = wb[sheet_name]
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row,
                                    min_col=1, max_col=ws.max_column):
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                        new_val = pattern.sub(shift_match, cell.value)
                        if new_val != cell.value:
                            cell.value = new_val
                            fixed_count += 1

        logger.info("cross_sheet_references_fixed", modified_sheet=modified_sheet, refs_updated=fixed_count)

    def capture_header_merges(self, ws, max_header_row=9) -> list:
        merges = []
        for m in ws.merged_cells.ranges:
            if m.min_row <= max_header_row:
                merges.append({
                    "min_row": m.min_row,
                    "max_row": m.max_row,
                    "min_col": m.min_col,
                    "max_col": m.max_col,
                })
        return merges

    def fix_header_merges(self, ws, original_merges: list, insert_col: int, insert_count: int):
        for m in list(ws.merged_cells.ranges):
            if m.min_row <= 9:
                try:
                    ws.unmerge_cells(str(m))
                except (KeyError, ValueError):
                    pass

        for m in original_merges:
            min_col = m["min_col"]
            max_col = m["max_col"]
            if min_col >= insert_col:
                min_col += insert_count
                max_col += insert_count
            try:
                ws.merge_cells(
                    start_row=m["min_row"],
                    start_column=min_col,
                    end_row=m["max_row"],
                    end_column=max_col,
                )
            except (ValueError, KeyError):
                pass

    def parse_period_date(self, period_str: str):
        try:
            return datetime.strptime(period_str, "%B %d, %Y")
        except ValueError:
            return period_str

    _parse_period_date = parse_period_date

    def _restore_external_links(self, original_path: str, output_path: str):
        external_files = []
        with zipfile.ZipFile(original_path, "r") as orig_zip:
            for name in orig_zip.namelist():
                if "externalLink" in name:
                    external_files.append(name)

        if not external_files:
            return

        temp_path = output_path + ".tmp"
        with zipfile.ZipFile(output_path, "r") as out_zip:
            with zipfile.ZipFile(temp_path, "w", zipfile.ZIP_DEFLATED) as new_zip:
                for item in out_zip.namelist():
                    if item in external_files:
                        continue
                    new_zip.writestr(item, out_zip.read(item))

                with zipfile.ZipFile(original_path, "r") as orig_zip:
                    for ef in external_files:
                        new_zip.writestr(ef, orig_zip.read(ef))

        shutil.move(temp_path, output_path)
        logger.info("external_links_restored", files=len(external_files))
