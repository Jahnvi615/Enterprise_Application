import re
from openpyxl.utils import get_column_letter
import structlog

logger = structlog.get_logger()

CASH_FLOW_SHEET_NAME = "Cash Flow"
INSERT_COL = 2
INSERT_COUNT = 1

SECTION_ROW_RANGES = [
    (9, 52, "operating activities"),
    (56, 73, "investing activities"),
    (77, 104, "financing activities"),
    (108, 109, "other"),
]


class CashFlowHandler:
    def __init__(self, orchestrator):
        self._orch = orchestrator

    def handle(self, wb, extraction_data: dict, context: dict):
        ws = wb[CASH_FLOW_SHEET_NAME]
        period = context["period"]

        values = self._prepare_data(extraction_data)

        logger.info(
            "cash_flow_handler_started",
            period=period,
            categories=len(values),
        )

        header_merges = self._orch.capture_header_merges(ws)

        ws.insert_cols(INSERT_COL, amount=INSERT_COUNT)
        self._orch.fix_shifted_formulas(ws, INSERT_COL, INSERT_COUNT)
        self._orch.fix_header_merges(ws, header_merges, INSERT_COL, INSERT_COUNT)

        self._style_new_column(ws)
        self._replicate_formulas(ws)
        self._set_analysis_header(ws)
        populated = self._populate_values(ws, values)

        logger.info("cash_flow_handler_completed", rows_populated=populated)

    def _prepare_data(self, extraction_data: dict) -> dict:
        periods = extraction_data.get("periods", [])
        rows = extraction_data.get("rows", [])

        if not periods:
            return {}

        most_recent_period = periods[0]
        values = {}

        for row in rows:
            if row.get("is_non_mappable"):
                continue
            category = row.get("mapped_category", "")
            if not category or category == "UNMAPPED":
                continue
            section = row.get("section", "").lower()
            value = row.get("values", {}).get(most_recent_period)
            if value is not None:
                values[(section, category)] = value

        return values

    def _style_new_column(self, ws):
        src_col = INSERT_COL + INSERT_COUNT
        for row_idx in range(1, ws.max_row + 1):
            self._orch.copy_cell_style(ws, row_idx, src_col, row_idx, INSERT_COL)

    def _replicate_formulas(self, ws):
        src_col = INSERT_COL + INSERT_COUNT
        src_letter = get_column_letter(src_col)
        new_letter = get_column_letter(INSERT_COL)
        count = 0

        for row_idx in range(1, ws.max_row + 1):
            src_cell = ws.cell(row=row_idx, column=src_col)
            if (
                src_cell.value
                and isinstance(src_cell.value, str)
                and src_cell.value.startswith("=")
            ):
                new_formula = re.sub(
                    rf"\b{re.escape(src_letter)}(?=\d)",
                    new_letter,
                    src_cell.value,
                )
                ws.cell(row=row_idx, column=INSERT_COL, value=new_formula)
                count += 1

        logger.info("cf_formulas_replicated", count=count)

    def _set_analysis_header(self, ws):
        ws.cell(row=8, column=INSERT_COL, value="As Allowed")

    def _populate_values(self, ws, values: dict) -> int:
        populated = 0
        src_col = INSERT_COL + INSERT_COUNT

        for range_start, range_end, section_key in SECTION_ROW_RANGES:
            for row_idx in range(range_start, range_end + 1):
                label = ws.cell(row=row_idx, column=1).value
                if not label:
                    continue

                label_clean = str(label).strip()

                # Only write to rows that are data rows (have numeric values in
                # adjacent column). Sub-headers like "(Increase) decrease in:"
                # have None in all value columns — leave those blank.
                adjacent_val = ws.cell(row=row_idx, column=src_col).value
                is_data_row = adjacent_val is not None and not (
                    isinstance(adjacent_val, str) and adjacent_val.startswith("=")
                )
                if not is_data_row:
                    continue

                key = (section_key, label_clean)
                if key in values:
                    ws.cell(row=row_idx, column=INSERT_COL, value=values[key] * 1000)
                    populated += 1
                    logger.debug(
                        "cf_value_populated",
                        row=row_idx,
                        section=section_key,
                        label=label_clean,
                        value=values[key],
                    )
                else:
                    ws.cell(row=row_idx, column=INSERT_COL, value=0)

        return populated
