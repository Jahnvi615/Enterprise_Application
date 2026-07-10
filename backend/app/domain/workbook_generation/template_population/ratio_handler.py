import copy
import re
from openpyxl.utils import get_column_letter
import structlog

logger = structlog.get_logger()

RATIO_SHEET_NAME = "Ratio"
BALANCE_SHEET_TREND_NAME = "Balance Sheet Trend"
INSERT_COL = 4
INSERT_COUNT = 1
HEADER_END_ROW = 8
DATA_START_ROW = 9


class RatioHandler:
    def __init__(self, orchestrator):
        self._orch = orchestrator

    def handle(self, wb, context: dict):
        if RATIO_SHEET_NAME not in wb.sheetnames:
            return

        ws = wb[RATIO_SHEET_NAME]

        logger.info("ratio_handler_started")

        all_merges = self._capture_all_merges(ws)
        header_cells = self._capture_header_cells(ws)

        ws.insert_cols(INSERT_COL, amount=INSERT_COUNT)
        self._orch.fix_shifted_formulas(ws, INSERT_COL, INSERT_COUNT)

        self._unmerge_all(ws)
        self._restore_header_cells(ws, header_cells)
        self._restore_all_merges(ws, all_merges)

        self._style_new_column(ws)
        self._set_header(ws)
        self._populate_formulas(ws)

        logger.info("ratio_handler_completed")

    # --- Capture Methods (before insertion) ---

    def _capture_all_merges(self, ws) -> list:
        merges = []
        for m in ws.merged_cells.ranges:
            merges.append({
                "min_row": m.min_row,
                "max_row": m.max_row,
                "min_col": m.min_col,
                "max_col": m.max_col,
            })
        return merges

    def _capture_header_cells(self, ws) -> dict:
        cells = {}
        for row_idx in range(1, HEADER_END_ROW + 1):
            for col_idx in range(INSERT_COL, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cells[(row_idx, col_idx)] = {
                    "value": cell.value,
                    "font": copy.copy(cell.font),
                    "fill": copy.copy(cell.fill),
                    "border": copy.copy(cell.border),
                    "alignment": copy.copy(cell.alignment),
                    "number_format": cell.number_format,
                }
        return cells

    # --- Restore Methods (after insertion) ---

    def _unmerge_all(self, ws):
        for m in list(ws.merged_cells.ranges):
            try:
                ws.unmerge_cells(str(m))
            except (KeyError, ValueError):
                pass

    def _restore_all_merges(self, ws, original_merges: list):
        for m in original_merges:
            min_col = m["min_col"]
            max_col = m["max_col"]
            if min_col >= INSERT_COL:
                min_col += INSERT_COUNT
                max_col += INSERT_COUNT
            try:
                ws.merge_cells(
                    start_row=m["min_row"],
                    start_column=min_col,
                    end_row=m["max_row"],
                    end_column=max_col,
                )
            except (ValueError, KeyError):
                pass

    def _restore_header_cells(self, ws, cells: dict):
        from openpyxl.cell.cell import MergedCell
        for (row_idx, orig_col), style in cells.items():
            new_col = orig_col + INSERT_COUNT
            cell = ws.cell(row=row_idx, column=new_col)
            if isinstance(cell, MergedCell):
                continue
            cell.value = style["value"]
            cell.font = style["font"]
            cell.fill = style["fill"]
            cell.border = style["border"]
            cell.alignment = style["alignment"]
            cell.number_format = style["number_format"]

    # --- New Column Setup ---

    def _style_new_column(self, ws):
        src_col = INSERT_COL + INSERT_COUNT
        for row_idx in range(1, ws.max_row + 1):
            self._orch.copy_cell_style(ws, row_idx, src_col, row_idx, INSERT_COL)

    def _set_header(self, ws):
        trend_new_col_letter = get_column_letter(2)
        ws.cell(row=5, column=INSERT_COL, value=f"='{BALANCE_SHEET_TREND_NAME}'!{trend_new_col_letter}6")
        ws.cell(row=6, column=INSERT_COL, value=f"='{BALANCE_SHEET_TREND_NAME}'!{trend_new_col_letter}7")
        ws.cell(row=7, column=INSERT_COL, value=f"='{BALANCE_SHEET_TREND_NAME}'!{trend_new_col_letter}8")
        ws.cell(row=8, column=INSERT_COL, value="As Allowed")

    def _populate_formulas(self, ws):
        src_col = INSERT_COL + INSERT_COUNT
        trend_existing_letter = get_column_letter(3)
        trend_new_letter = get_column_letter(2)

        populated = 0
        for row_idx in range(DATA_START_ROW, ws.max_row + 1):
            src_cell = ws.cell(row=row_idx, column=src_col)
            if not src_cell.value or not isinstance(src_cell.value, str) or not src_cell.value.startswith("="):
                continue

            new_formula = self._adapt_formula(src_cell.value, trend_existing_letter, trend_new_letter)
            ws.cell(row=row_idx, column=INSERT_COL, value=new_formula)
            populated += 1

        logger.info("ratio_formulas_populated", count=populated)

    def _adapt_formula(self, formula: str, trend_src_letter: str, trend_new_letter: str) -> str:
        trend_pattern = re.compile(
            r"('{0}'!)({1})(\d+)".format(re.escape(BALANCE_SHEET_TREND_NAME), trend_src_letter)
        )
        result = trend_pattern.sub(
            lambda m: f"{m.group(1)}{trend_new_letter}{m.group(3)}",
            formula,
        )
        pnl_pattern = re.compile(
            r"('P&L'!)({0})(\d+)".format(trend_src_letter)
        )
        return pnl_pattern.sub(
            lambda m: f"{m.group(1)}{trend_new_letter}{m.group(3)}",
            result,
        )
