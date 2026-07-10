import re
from openpyxl.utils import get_column_letter
import structlog

logger = structlog.get_logger()

PNL_SHEET_NAME = "P&L"
INSERT_COL = 2
INSERT_COUNT = 1
DATA_ROW_START = 9
ANALYSIS_HEADER_ROW = 8
ANALYSIS_HEADER_VALUE = "As Given"

# Template stores these as positive values and subtracts them in formulas.
# PDF shows them in parentheses (negative), so we flip to positive here.
ALWAYS_POSITIVE_CATEGORIES = {
    "Cost of Contract Revenues Earned",
    "Selling general administrative expenses",
}


class PAndLHandler:
    def __init__(self, orchestrator):
        self._orch = orchestrator

    def handle(self, wb, extraction_data: dict, context: dict):
        if PNL_SHEET_NAME not in wb.sheetnames:
            logger.warning("pnl_sheet_not_found", available=wb.sheetnames)
            return

        ws = wb[PNL_SHEET_NAME]
        period = context["period"]
        values = self._prepare_data(extraction_data)

        logger.info(
            "pnl_handler_started",
            period=period,
            categories=len(values),
        )

        header_merges = self._orch.capture_header_merges(ws)

        # Must unmerge BEFORE insert_cols — openpyxl silently drops sub-cell
        # values during the column shift when a merge covers the destination cells.
        for m_range in list(ws.merged_cells.ranges):
            if m_range.min_row <= 9:
                try:
                    ws.unmerge_cells(str(m_range))
                except (KeyError, ValueError):
                    pass

        ws.insert_cols(INSERT_COL, amount=INSERT_COUNT)
        self._orch.fix_shifted_formulas(ws, INSERT_COL, INSERT_COUNT)
        self._orch.fix_header_merges(ws, header_merges, INSERT_COL, INSERT_COUNT)

        self._style_new_column(ws)
        self._replicate_formulas(ws)
        self._set_header(ws)
        populated = self._populate_values(ws, values)

        logger.info("pnl_handler_completed", rows_populated=populated)

    def _prepare_data(self, extraction_data: dict) -> dict:
        periods = extraction_data.get("periods", [])
        rows = extraction_data.get("rows", [])

        if not periods:
            return {}

        most_recent = periods[0]
        values = {}

        for row in rows:
            if row.get("is_non_mappable"):
                continue
            category = row.get("mapped_category", "")
            if not category or category.startswith("UNMAPPED"):
                continue
            value = row.get("values", {}).get(most_recent)
            if value is not None:
                values[category] = value

        return values

    def _style_new_column(self, ws):
        src_col = INSERT_COL + INSERT_COUNT
        for row_idx in range(1, ws.max_row + 1):
            self._orch.copy_cell_style(ws, row_idx, src_col, row_idx, INSERT_COL)

    def _replicate_formulas(self, ws):
        src_col = INSERT_COL + INSERT_COUNT
        src_letter = get_column_letter(src_col)
        new_letter = get_column_letter(INSERT_COL)
        count = 0

        for row_idx in range(1, ws.max_row + 1):
            src_cell = ws.cell(row=row_idx, column=src_col)
            if (
                src_cell.value
                and isinstance(src_cell.value, str)
                and src_cell.value.startswith("=")
            ):
                new_formula = re.sub(
                    rf"\b{re.escape(src_letter)}(?=\d)",
                    new_letter,
                    src_cell.value,
                )
                ws.cell(row=row_idx, column=INSERT_COL, value=new_formula)
                count += 1

        logger.info("pnl_formulas_replicated", count=count)

    def _set_header(self, ws):
        ws.cell(row=ANALYSIS_HEADER_ROW, column=INSERT_COL, value=ANALYSIS_HEADER_VALUE)

    def _populate_values(self, ws, values: dict) -> int:
        populated = 0

        for row_idx in range(DATA_ROW_START, ws.max_row + 1):
            label = ws.cell(row=row_idx, column=1).value
            if not label:
                continue

            label_clean = str(label).strip()

            if label_clean in values:
                extracted = values[label_clean] * 1000
                if label_clean in ALWAYS_POSITIVE_CATEGORIES and extracted < 0:
                    extracted = abs(extracted)
                ws.cell(row=row_idx, column=INSERT_COL, value=extracted)
                populated += 1
                logger.debug(
                    "pnl_value_populated",
                    row=row_idx,
                    label=label_clean,
                    value=values[label_clean],
                )

        return populated
