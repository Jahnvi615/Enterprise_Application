from openpyxl.utils import get_column_letter
import structlog

logger = structlog.get_logger()

BALANCE_SHEET_TREND_NAME = "Balance Sheet Trend"
BALANCE_SHEET_NAME = "Balance Sheet"
DATA_START_ROW = 11
INSERT_COL = 2
INSERT_COUNT = 1


class BalanceSheetTrendHandler:
    def __init__(self, orchestrator):
        self._orch = orchestrator

    def handle(self, wb, context: dict):
        ws = wb[BALANCE_SHEET_TREND_NAME]

        bs_as_allowed_col = context["bs_as_allowed_col"]
        bs_as_allowed_letter = get_column_letter(bs_as_allowed_col)

        logger.info("balance_sheet_trend_handler_started")

        header_merges = self._orch.capture_header_merges(ws)

        ws.insert_cols(INSERT_COL, amount=INSERT_COUNT)
        self._orch.fix_shifted_formulas(ws, INSERT_COL, INSERT_COUNT)
        self._orch.fix_header_merges(ws, header_merges, INSERT_COL, INSERT_COUNT)

        self._fix_column_widths(ws)
        self._style_new_column(ws)
        self._set_header(ws, context["insert_col"])
        self._populate_from_balance_sheet(ws, bs_as_allowed_letter)
        self._propagate_total_formulas(ws)

        logger.info("balance_sheet_trend_handler_completed")

    def _fix_column_widths(self, ws):
        ref_letter = get_column_letter(INSERT_COL + INSERT_COUNT)
        ref_width = 13
        if ref_letter in ws.column_dimensions and ws.column_dimensions[ref_letter].width:
            ref_width = ws.column_dimensions[ref_letter].width

        for col_idx in range(INSERT_COL, ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            if col_letter not in ws.column_dimensions or not ws.column_dimensions[col_letter].width or ws.column_dimensions[col_letter].width < ref_width:
                ws.column_dimensions[col_letter].width = ref_width

    def _style_new_column(self, ws):
        src_col = INSERT_COL + INSERT_COUNT

        for row_idx in range(1, ws.max_row + 1):
            self._orch.copy_cell_style(ws, row_idx, src_col, row_idx, INSERT_COL)

    def _set_header(self, ws, bs_as_given_col: int):
        bs_as_given_letter = get_column_letter(bs_as_given_col)
        ws.cell(row=6, column=INSERT_COL, value=f"='{BALANCE_SHEET_NAME}'!{bs_as_given_letter}6")
        ws.cell(row=7, column=INSERT_COL, value=f"='{BALANCE_SHEET_NAME}'!{bs_as_given_letter}7")
        ws.cell(row=8, column=INSERT_COL, value=f"='{BALANCE_SHEET_NAME}'!{bs_as_given_letter}8")
        ws.cell(row=9, column=INSERT_COL, value="As Allowed")

    def _populate_from_balance_sheet(self, ws, bs_as_allowed_letter: str):
        populated = 0

        for row_idx in range(DATA_START_ROW, ws.max_row + 1):
            label = ws.cell(row=row_idx, column=1).value
            if not label:
                continue

            label_clean = str(label).strip()
            if label_clean.lower().startswith("total") or label_clean in ("Working Capital", "Aggregate Program", "WC % Case", "NW % Case"):
                continue

            if label_clean in ("Current Assets", "Non Current Assets", "Current Liabilities",
                               "Non-Current Liabilities", "Equity", "Commitments & Contingencies"):
                continue

            ws.cell(
                row=row_idx,
                column=INSERT_COL,
                value=f"='{BALANCE_SHEET_NAME}'!{bs_as_allowed_letter}{row_idx}",
            )
            populated += 1

        logger.info("trend_values_populated", count=populated)

    def _propagate_total_formulas(self, ws):
        src_col = INSERT_COL + INSERT_COUNT
        new_col_letter = get_column_letter(INSERT_COL)
        src_col_letter = get_column_letter(src_col)

        total_count = 0
        for row_idx in range(DATA_START_ROW, ws.max_row + 1):
            label = ws.cell(row=row_idx, column=1).value
            if not label:
                continue

            label_clean = str(label).strip()
            if not (label_clean.lower().startswith("total") or
                    label_clean in ("Working Capital", "Aggregate Program", "WC % Case", "NW % Case")):
                continue

            src_cell = ws.cell(row=row_idx, column=src_col)
            if src_cell.value and isinstance(src_cell.value, str) and src_cell.value.startswith("="):
                new_formula = src_cell.value.replace(src_col_letter, new_col_letter)
                ws.cell(row=row_idx, column=INSERT_COL, value=new_formula)
                total_count += 1

        logger.info("trend_total_formulas_propagated", count=total_count)
