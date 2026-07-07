from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import structlog

logger = structlog.get_logger()

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
SECTION_FONT = Font(name="Calibri", size=10, bold=True, color="1F4E79")
SECTION_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
TOTAL_FONT = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
TOTAL_FILL = PatternFill(start_color="2B6CB0", end_color="2B6CB0", fill_type="solid")
DATA_FONT = Font(name="Calibri", size=10)
UNMAPPED_FONT = Font(name="Calibri", size=10, color="E53E3E")

BORDER_BOTTOM = Border(bottom=Side(style="thin", color="D9D9D9"))
BORDER_TOTAL = Border(
    top=Side(style="thin", color="1F4E79"),
    bottom=Side(style="thin", color="1F4E79"),
)

NUMBER_FORMAT = '#,##0;(#,##0)'


def generate_sample_workbook(extraction_results: dict, output_path: str) -> str:
    wb = Workbook()
    wb.remove(wb.active)

    for statement_type, data in extraction_results.items():
        sheet_name = _get_sheet_name(statement_type)
        has_normalized_label = any(r.get("normalized_label") for r in data.get("rows", []))
        has_spreading = any("spreading_rule" in r for r in data.get("rows", []))
        _create_sheet(wb, sheet_name, data, has_normalized_label, has_spreading)

    wb.save(output_path)
    logger.info("sample_workbook_generated", path=output_path)
    return output_path


def _get_sheet_name(statement_type: str) -> str:
    names = {
        "balance_sheet": "BalanceSheet",
        "cash_flow": "CashFlow",
    }
    return names.get(statement_type, statement_type)


def _create_sheet(wb: Workbook, sheet_name: str, data: dict, has_normalized_label: bool = False, has_spreading: bool = False):
    ws = wb.create_sheet(title=sheet_name)
    periods = data.get("periods", [])
    rows = data.get("rows", [])

    headers = ["Source Label"]
    if has_normalized_label:
        headers.append("Normalized Label")
    headers += ["Section", "Row Type", "Is Non-Mappable", "Mapped Category"]
    if has_spreading:
        headers.append("Spreading Rule")
    headers += periods

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.row_dimensions[1].height = 28

    period_start_col = len(headers) - len(periods) + 1

    for row_idx, row_data in enumerate(rows, 2):
        row_type = row_data["row_type"]
        is_section = row_type == "section_header"
        is_total = row_type == "total"

        col = 1
        ws.cell(row=row_idx, column=col, value=row_data["source_label"])
        col += 1

        if has_normalized_label:
            ws.cell(row=row_idx, column=col, value=row_data.get("normalized_label", ""))
            col += 1

        ws.cell(row=row_idx, column=col, value=row_data["section"])
        col += 1
        ws.cell(row=row_idx, column=col, value=row_type)
        col += 1
        ws.cell(row=row_idx, column=col, value=row_data["is_non_mappable"])
        col += 1

        mapped_category = row_data.get("mapped_category", "")
        mapped_cell = ws.cell(row=row_idx, column=col, value=mapped_category)
        col += 1

        if has_spreading:
            ws.cell(row=row_idx, column=col, value=row_data.get("spreading_rule", ""))
            col += 1

        for period_idx, period in enumerate(periods):
            value = row_data["values"].get(period)
            val_cell = ws.cell(row=row_idx, column=period_start_col + period_idx, value=value)
            if value is not None:
                val_cell.number_format = NUMBER_FORMAT
                val_cell.alignment = Alignment(horizontal="right")

        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=c)
            if is_section:
                cell.font = SECTION_FONT
                cell.fill = SECTION_FILL
                cell.border = BORDER_BOTTOM
            elif is_total:
                cell.font = TOTAL_FONT
                cell.fill = TOTAL_FILL
                cell.border = BORDER_TOTAL
            else:
                cell.font = DATA_FONT
                cell.border = BORDER_BOTTOM

        if not is_section and not is_total:
            ws.cell(row=row_idx, column=1).alignment = Alignment(indent=2)

        if mapped_category == "UNMAPPED" and not is_section and not is_total:
            mapped_cell.font = UNMAPPED_FONT

    col_idx = 1
    ws.column_dimensions[get_column_letter(col_idx)].width = 65
    col_idx += 1
    if has_normalized_label:
        ws.column_dimensions[get_column_letter(col_idx)].width = 45
        col_idx += 1
    ws.column_dimensions[get_column_letter(col_idx)].width = 25
    col_idx += 1
    ws.column_dimensions[get_column_letter(col_idx)].width = 16
    col_idx += 1
    ws.column_dimensions[get_column_letter(col_idx)].width = 16
    col_idx += 1
    ws.column_dimensions[get_column_letter(col_idx)].width = 30
    col_idx += 1
    if has_spreading:
        ws.column_dimensions[get_column_letter(col_idx)].width = 45
        col_idx += 1
    for i in range(len(periods)):
        ws.column_dimensions[get_column_letter(col_idx + i)].width = 24

    ws.freeze_panes = "A2"
    ws.sheet_properties.tabColor = "1F4E79"
