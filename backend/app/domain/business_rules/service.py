import json
from pathlib import Path
from openpyxl.utils import get_column_letter
import structlog

logger = structlog.get_logger()

CONFIG_DIR = Path(__file__).parent / "config"

SECTION_NORMALIZE = {
    "current assets": "current_assets",
    "non current assets": "noncurrent_assets",
    "noncurrent assets": "noncurrent_assets",
    "non-current assets": "noncurrent_assets",
    "current liabilities": "current_liabilities",
    "non-current liabilities": "noncurrent_liabilities",
    "noncurrent liabilities": "noncurrent_liabilities",
    "non current liabilities": "noncurrent_liabilities",
    "equity": "equity",
    "commitments and contingencies": "commitments",
    "commitments & contingencies": "commitments",
}

DATA_START_ROW = 11


class SpreadingRulesService:
    def __init__(self):
        self._rules = self._load_rules("balance_sheet_spreading_rules.json")

    def apply(self, ws, as_given_col: int, as_allowed_col: int, remarks_col: int):
        as_given_letter = get_column_letter(as_given_col)
        current_section = ""
        applied = 0

        for row_idx in range(DATA_START_ROW, ws.max_row + 1):
            label = ws.cell(row=row_idx, column=1).value
            if not label:
                continue

            label_clean = str(label).strip()
            norm = self._normalize_section(label_clean)
            if norm in SECTION_NORMALIZE.values():
                current_section = norm
                continue

            if label_clean.lower().startswith("total"):
                continue

            as_given_value = ws.cell(row=row_idx, column=as_given_col).value
            if as_given_value is None:
                continue

            rule = self._find_rule(current_section, label_clean)
            if not rule:
                continue

            rule_type = rule["rule_type"]
            multiplier = rule["multiplier"]
            rule_text = rule["rule_text"]

            if rule_type == "no_changes":
                ws.cell(
                    row=row_idx,
                    column=as_allowed_col,
                    value=f"={as_given_letter}{row_idx}*1",
                )
                if rule_text.lower() != "no changes":
                    ws.cell(row=row_idx, column=remarks_col, value=rule_text)

            elif rule_type in ("percentage_deferred", "fully_disallowed", "fully_moved", "reclassified"):
                ws.cell(
                    row=row_idx,
                    column=as_allowed_col,
                    value=f"={as_given_letter}{row_idx}*{multiplier}",
                )
                if rule_text.lower() != "no changes":
                    ws.cell(row=row_idx, column=remarks_col, value=rule_text)

            elif rule_type == "special":
                ws.cell(
                    row=row_idx,
                    column=as_allowed_col,
                    value=f"={as_given_letter}{row_idx}*{multiplier}",
                )
                if rule_text.lower() != "no changes":
                    ws.cell(row=row_idx, column=remarks_col, value=rule_text)

            else:
                ws.cell(
                    row=row_idx,
                    column=as_allowed_col,
                    value=f"={as_given_letter}{row_idx}*1",
                )

            applied += 1

        logger.info("spreading_rules_applied", rules_applied=applied)

    def enrich_extraction(self, extraction_data: dict) -> dict:
        rows = extraction_data.get("rows", [])
        for row in rows:
            if row.get("is_non_mappable"):
                row["spreading_rule"] = ""
                continue
            category = row.get("mapped_category", "")
            if not category or category == "UNMAPPED":
                row["spreading_rule"] = ""
                continue
            section = row.get("section", "")
            norm_section = self._normalize_section(section)
            rule = self._find_rule(norm_section, category)
            row["spreading_rule"] = rule["rule_text"] if rule else ""
        return extraction_data

    def _find_rule(self, section: str, label: str) -> dict | None:
        section_rules = self._rules.get(section, {})
        if label in section_rules:
            return section_rules[label]
        return None

    def _normalize_section(self, section: str) -> str:
        return SECTION_NORMALIZE.get(section.lower().strip(), section.lower().strip())

    def _load_rules(self, filename: str) -> dict:
        path = CONFIG_DIR / filename
        if not path.exists():
            logger.error("spreading_rules_not_found", path=str(path))
            return {}
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
        total_rules = sum(len(v) for v in data.values())
        logger.info("spreading_rules_loaded", sections=len(data), total_rules=total_rules)
        return data
