from openpyxl.utils import get_column_letter
import structlog

logger = structlog.get_logger()

_ROW_ACCOUNTS_RECEIVABLE = 15
_ROW_AR_RELATED_PARTIES = 20
_ROW_COSTS_EARNINGS_EXCESS = 16
_ROW_NONCURRENT_MARKETABLE_SECURITIES = 30
_ROW_GOODWILL = 36
_ROW_INTANGIBLES = 37
_ROW_INVESTMENT_OTHER = 42
_ROW_BILLINGS_IN_EXCESS = 54
_ROW_ACCRUED_EXPENSES = 55
_ROW_LONG_TERM_DEBT = 65
_ROW_DEFERRED_COMP = 66
_ROW_OTHER_LONG_TERM_LIABILITIES = 68
_ROW_DUE_TO_JV_NCI = 69
_ROW_LLC_MEMBERS_EQUITY = 82
_ROW_UNDERWRITING_ADJUSTMENTS = 85


class BalanceSheetCustomRulesService:

    def apply(
        self,
        ws,
        as_given_col: int,
        as_allowed_col: int,
        remarks_col: int,
        notes_data: dict,
    ) -> None:
        as_given_letter = get_column_letter(as_given_col)
        applied = []

        if self._apply_accounts_receivable_jv_adjustment(
            ws, as_given_col, as_given_letter, as_allowed_col, remarks_col, notes_data
        ):
            applied.append("accounts_receivable_jv")

        if self._apply_costs_earnings_excess_remark(ws, remarks_col, notes_data):
            applied.append("costs_earnings_excess_remark")

        if self._apply_noncurrent_marketable_securities(ws, as_given_letter, as_allowed_col):
            applied.append("noncurrent_marketable_securities")

        if self._apply_investment_other_remark(ws, remarks_col):
            applied.append("investment_other_remark")

        if self._apply_billings_in_excess_remark(ws, remarks_col):
            applied.append("billings_in_excess_remark")

        if self._apply_accrued_expenses(ws, as_given_col, remarks_col, notes_data):
            applied.append("accrued_expenses")

        if self._apply_long_term_debt_remark(ws, remarks_col):
            applied.append("long_term_debt_remark")

        if self._apply_other_long_term_liabilities(ws, as_given_col, as_given_letter, as_allowed_col, remarks_col):
            applied.append("other_long_term_liabilities")

        if self._apply_due_to_jv_nci(ws, as_given_col, as_given_letter, as_allowed_col):
            applied.append("due_to_jv_nci")

        if self._apply_llc_members_equity_remark(ws, remarks_col):
            applied.append("llc_members_equity_remark")

        if self._apply_underwriting_adjustment(
            ws, as_given_col, as_given_letter, as_allowed_col, remarks_col, notes_data
        ):
            applied.append("underwriting_adjustment")

        logger.info("bs_custom_rules_applied", adjustments=applied)

    def _apply_accounts_receivable_jv_adjustment(
        self,
        ws,
        as_given_col: int,
        as_given_letter: str,
        as_allowed_col: int,
        remarks_col: int,
        notes_data: dict,
    ) -> bool:
        jv_ar_amount = notes_data.get("jv_ar_amount")
        if not jv_ar_amount:
            logger.debug("bs_custom_ar_jv_skipped", reason="jv_ar_amount_not_available")
            return False

        # Row 15: deduct JV-related A/R from As Allowed (As Given must be populated)
        as_given_value = ws.cell(row=_ROW_ACCOUNTS_RECEIVABLE, column=as_given_col).value
        if as_given_value is not None:
            ws.cell(
                row=_ROW_ACCOUNTS_RECEIVABLE,
                column=as_allowed_col,
                value=f"={as_given_letter}{_ROW_ACCOUNTS_RECEIVABLE}*1-{jv_ar_amount}",
            )
            ws.cell(
                row=_ROW_ACCOUNTS_RECEIVABLE,
                column=remarks_col,
                value=(
                    f"Includes Accounts receivable, net ${int(as_given_value):,} - "
                    f"Work performed for unconsolidated partnerships and "
                    f"joint ventures ${jv_ar_amount:,}"
                ),
            )
            logger.debug(
                "bs_custom_ar_written",
                row=_ROW_ACCOUNTS_RECEIVABLE,
                as_given=as_given_value,
                jv_amount=jv_ar_amount,
            )

        # Row 20: As Given is always empty — written unconditionally to show disallowance
        ws.cell(
            row=_ROW_AR_RELATED_PARTIES,
            column=as_allowed_col,
            value=f"={jv_ar_amount}*0",
        )
        ws.cell(
            row=_ROW_AR_RELATED_PARTIES,
            column=remarks_col,
            value=(
                "Includes Work performed for unconsolidated partnerships and "
                "joint ventures (100% Disallowed from asset and equity)"
            ),
        )
        logger.debug(
            "bs_custom_ar_related_parties_written",
            row=_ROW_AR_RELATED_PARTIES,
            jv_amount=jv_ar_amount,
        )

        return True

    def _apply_costs_earnings_excess_remark(
        self,
        ws,
        remarks_col: int,
        notes_data: dict,
    ) -> bool:
        unbilled_amount = notes_data.get("unbilled_receivables")
        if not unbilled_amount:
            logger.debug("bs_custom_costs_earnings_remark_skipped", reason="unbilled_receivables_not_available")
            return False

        ws.cell(
            row=_ROW_COSTS_EARNINGS_EXCESS,
            column=remarks_col,
            value=f"Includes Contract Assets (Includes Unbilled receivables ${unbilled_amount:,})",
        )
        logger.debug("bs_custom_costs_earnings_remark_written", row=_ROW_COSTS_EARNINGS_EXCESS, unbilled=unbilled_amount)
        return True

    def _apply_noncurrent_marketable_securities(
        self,
        ws,
        as_given_letter: str,
        as_allowed_col: int,
    ) -> bool:
        # As Given is always empty for this row — formula references current Marketable
        # Securities (row 14) As Given column directly, deferred at 20%
        ws.cell(
            row=_ROW_NONCURRENT_MARKETABLE_SECURITIES,
            column=as_allowed_col,
            value=f"={as_given_letter}14*0.2",
        )
        logger.debug("bs_custom_noncurrent_mkt_sec_written", row=_ROW_NONCURRENT_MARKETABLE_SECURITIES)
        return True

    def _apply_investment_other_remark(
        self,
        ws,
        remarks_col: int,
    ) -> bool:
        ws.cell(
            row=_ROW_INVESTMENT_OTHER,
            column=remarks_col,
            value="Includes Investments",
        )
        logger.debug("bs_custom_investment_other_remark_written", row=_ROW_INVESTMENT_OTHER)
        return True

    def _apply_billings_in_excess_remark(
        self,
        ws,
        remarks_col: int,
    ) -> bool:
        ws.cell(
            row=_ROW_BILLINGS_IN_EXCESS,
            column=remarks_col,
            value="Includes Contract liabilities",
        )
        logger.debug("bs_custom_billings_in_excess_remark_written", row=_ROW_BILLINGS_IN_EXCESS)
        return True

    def _apply_accrued_expenses(
        self,
        ws,
        as_given_col: int,
        remarks_col: int,
        notes_data: dict,
    ) -> bool:
        accrued_salaries = notes_data.get("accrued_salaries_wages")
        other_accrued = notes_data.get("other_accrued_liabilities")
        if accrued_salaries is None or other_accrued is None:
            logger.debug(
                "bs_custom_accrued_expenses_skipped",
                reason="accrued_components_not_available",
                accrued_salaries=accrued_salaries,
                other_accrued=other_accrued,
            )
            return False

        ws.cell(
            row=_ROW_ACCRUED_EXPENSES,
            column=as_given_col,
            value=f"={accrued_salaries}+{other_accrued}",
        )
        ws.cell(
            row=_ROW_ACCRUED_EXPENSES,
            column=remarks_col,
            value=(
                f"Includes Accrued salaries, wages and benefits ${accrued_salaries:,} + "
                f"Other accrued liabilities ${other_accrued:,}"
            ),
        )
        logger.debug(
            "bs_custom_accrued_expenses_written",
            row=_ROW_ACCRUED_EXPENSES,
            accrued_salaries=accrued_salaries,
            other_accrued=other_accrued,
        )
        return True

    def _apply_long_term_debt_remark(
        self,
        ws,
        remarks_col: int,
    ) -> bool:
        ws.cell(row=_ROW_LONG_TERM_DEBT, column=remarks_col).value = None
        logger.debug("bs_custom_long_term_debt_remark_cleared", row=_ROW_LONG_TERM_DEBT)
        return True

    def _apply_other_long_term_liabilities(
        self,
        ws,
        as_given_col: int,
        as_given_letter: str,
        as_allowed_col: int,
        remarks_col: int,
    ) -> bool:
        dc_as_allowed = ws.cell(row=_ROW_DEFERRED_COMP, column=as_allowed_col).value
        if not dc_as_allowed:
            logger.debug("bs_custom_oll_skipped", reason="deferred_comp_as_allowed_not_present")
            return False

        # DC As Allowed can be a static number (DC has no As Given — SpreadingRulesService
        # skips it) or a formula string if DC was extracted and As Given is populated.
        # Read dc_value from As Allowed directly when it is a static number.
        if isinstance(dc_as_allowed, (int, float)):
            dc_value = dc_as_allowed
        else:
            dc_value = ws.cell(row=_ROW_DEFERRED_COMP, column=as_given_col).value
            if not dc_value:
                logger.debug("bs_custom_oll_skipped", reason="deferred_comp_value_not_available")
                return False

        oll_ag_value = ws.cell(row=_ROW_OTHER_LONG_TERM_LIABILITIES, column=as_given_col).value
        if oll_ag_value is None:
            logger.debug("bs_custom_oll_skipped", reason="oll_as_given_not_present")
            return False

        as_allowed_letter = get_column_letter(as_allowed_col)
        ws.cell(
            row=_ROW_OTHER_LONG_TERM_LIABILITIES,
            column=as_allowed_col,
            value=f"={as_given_letter}{_ROW_OTHER_LONG_TERM_LIABILITIES}*1-{as_allowed_letter}{_ROW_DEFERRED_COMP}",
        )
        ws.cell(
            row=_ROW_OTHER_LONG_TERM_LIABILITIES,
            column=remarks_col,
            value=(
                f"Includes Other non current liabilities ${int(oll_ag_value):,} - "
                f"Deferred compensation ${int(dc_value):,}"
            ),
        )
        logger.debug(
            "bs_custom_oll_written",
            row=_ROW_OTHER_LONG_TERM_LIABILITIES,
            oll_as_given=oll_ag_value,
            dc_value=dc_value,
        )
        return True

    def _apply_due_to_jv_nci(
        self,
        ws,
        as_given_col: int,
        as_given_letter: str,
        as_allowed_col: int,
    ) -> bool:
        llc_as_given = ws.cell(row=_ROW_LLC_MEMBERS_EQUITY, column=as_given_col).value
        if llc_as_given is None:
            logger.debug("bs_custom_jv_nci_skipped", reason="llc_members_equity_as_given_not_present")
            return False

        ws.cell(
            row=_ROW_DUE_TO_JV_NCI,
            column=as_allowed_col,
            value=f"={as_given_letter}{_ROW_LLC_MEMBERS_EQUITY}",
        )
        logger.debug("bs_custom_jv_nci_written", row=_ROW_DUE_TO_JV_NCI, llc_as_given=llc_as_given)
        return True

    def _apply_llc_members_equity_remark(
        self,
        ws,
        remarks_col: int,
    ) -> bool:
        ws.cell(
            row=_ROW_LLC_MEMBERS_EQUITY,
            column=remarks_col,
            value="Includes Non Controlling interest (100% Reclassified from equity to long term liability)",
        )
        logger.debug("bs_custom_llc_members_equity_remark_written", row=_ROW_LLC_MEMBERS_EQUITY)
        return True

    def _apply_underwriting_adjustment(
        self,
        ws,
        as_given_col: int,
        as_given_letter: str,
        as_allowed_col: int,
        remarks_col: int,
        notes_data: dict,
    ) -> bool:
        jv_ar_amount = notes_data.get("jv_ar_amount")
        if not jv_ar_amount:
            logger.debug("bs_custom_underwriting_skipped", reason="jv_ar_amount_not_available")
            return False

        goodwill_ag = ws.cell(row=_ROW_GOODWILL, column=as_given_col).value
        intangibles_ag = ws.cell(row=_ROW_INTANGIBLES, column=as_given_col).value

        if goodwill_ag and intangibles_ag:
            # Condition 3: Goodwill + Intangibles + JV amount all present
            ws.cell(
                row=_ROW_UNDERWRITING_ADJUSTMENTS,
                column=as_allowed_col,
                value=f"=-{as_given_letter}{_ROW_GOODWILL}-{as_given_letter}{_ROW_INTANGIBLES}-{jv_ar_amount}",
            )
            ws.cell(
                row=_ROW_UNDERWRITING_ADJUSTMENTS,
                column=remarks_col,
                value=(
                    f"Includes Goodwill ${int(goodwill_ag):,} + "
                    f"Intangible Assets ${int(intangibles_ag):,} + "
                    f"A/R related to Partnerships and Joint ventures ${jv_ar_amount:,} "
                    f"(100% Disallowed from asset and equity)"
                ),
            )
            logger.debug(
                "bs_custom_underwriting_written",
                row=_ROW_UNDERWRITING_ADJUSTMENTS,
                condition=3,
                goodwill=goodwill_ag,
                intangibles=intangibles_ag,
                jv_amount=jv_ar_amount,
            )

        elif goodwill_ag:
            # Condition 2: Goodwill + JV amount present, no Intangibles
            ws.cell(
                row=_ROW_UNDERWRITING_ADJUSTMENTS,
                column=as_allowed_col,
                value=f"=-{as_given_letter}{_ROW_GOODWILL}-{jv_ar_amount}",
            )
            ws.cell(
                row=_ROW_UNDERWRITING_ADJUSTMENTS,
                column=remarks_col,
                value=(
                    f"Includes Goodwill ${int(goodwill_ag):,} + "
                    f"A/R related to Partnerships and Joint ventures ${jv_ar_amount:,} "
                    f"(100% Disallowed from asset and equity)"
                ),
            )
            logger.debug(
                "bs_custom_underwriting_written",
                row=_ROW_UNDERWRITING_ADJUSTMENTS,
                condition=2,
                goodwill=goodwill_ag,
                jv_amount=jv_ar_amount,
            )

        else:
            # Condition 1: JV amount only, no Goodwill and no Intangibles
            ws.cell(
                row=_ROW_UNDERWRITING_ADJUSTMENTS,
                column=as_allowed_col,
                value=f"=-{jv_ar_amount}",
            )
            ws.cell(
                row=_ROW_UNDERWRITING_ADJUSTMENTS,
                column=remarks_col,
                value="Includes Work performed for unconsolidated partnerships and joint ventures (100% Disallowed from asset and equity)",
            )
            logger.debug(
                "bs_custom_underwriting_written",
                row=_ROW_UNDERWRITING_ADJUSTMENTS,
                condition=1,
                jv_amount=jv_ar_amount,
            )

        return True
